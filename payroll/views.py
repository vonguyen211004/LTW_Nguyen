from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Q
from django.http import HttpResponse
from django.utils import timezone
from django.db import transaction
from decimal import Decimal
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Payroll, PayrollDetail, PayrollAllowance, PayrollDeduction
from employees.models import Employee, Position
from .forms import PayrollForm, PayrollDetailForm


@login_required
def payroll_list(request):
	month_filter = request.GET.get('month', '')
	year_filter = request.GET.get('year', '')
	position_filter = request.GET.get('position', '')
	status_filter = request.GET.get('status', '')

	# Lấy tháng và năm hiện tại nếu không có filter
	today = timezone.now()
	if not month_filter:
		month_filter = today.month
	if not year_filter:
		year_filter = today.year

	payrolls = Payroll.objects.all()

	# Kiểm tra quyền admin
	is_admin = request.user.is_superuser or request.user.groups.filter(name='HR Admin').exists()

	# Nếu không phải admin, chỉ hiển thị bảng lương không bị vô hiệu hóa
	if not is_admin:
		payrolls = payrolls.exclude(status='disabled')

	# Lọc theo tháng và năm
	if month_filter and year_filter:
		payrolls = payrolls.filter(month=month_filter, year=year_filter)

	# Lọc theo vị trí
	if position_filter:
		payrolls = payrolls.filter(position_id=position_filter)

	# Lọc theo trạng thái
	if status_filter:
		payrolls = payrolls.filter(status=status_filter)

	positions = Position.objects.all()

	return render(request, 'payroll/payroll_list.html', {
		'payrolls': payrolls,
		'positions': positions,
		'month_filter': int(month_filter) if month_filter else '',
		'year_filter': int(year_filter) if year_filter else '',
		'position_filter': position_filter,
		'status_filter': status_filter,
		'current_month': today.month,
		'current_year': today.year,
		'is_admin': is_admin
	})


@login_required
def payroll_detail(request, pk):
	payroll = get_object_or_404(Payroll, pk=pk)

	# Kiểm tra quyền admin
	is_admin = request.user.is_superuser or request.user.groups.filter(name='HR Admin').exists()

	# Nếu không phải admin và bảng lương bị vô hiệu hóa, chuyển hướng về danh sách
	if not is_admin and payroll.status == 'disabled':
		messages.warning(request, "Bảng lương này đã bị vô hiệu hóa.")
		return redirect('payroll_list')

	# Lấy chi tiết lương
	details = payroll.payroll_details.all()

	# Tính tổng
	aggregate = {
		'sum_gross_salary': details.aggregate(Sum('gross_salary'))['gross_salary__sum'] or 0,
		'sum_income_tax': details.aggregate(Sum('income_tax'))['income_tax__sum'] or 0,
		'sum_discipline': details.aggregate(Sum('discipline_amount'))['discipline_amount__sum'] or 0,
		'sum_reward': details.aggregate(Sum('reward_amount'))['reward_amount__sum'] or 0,
		'sum_deduction': details.aggregate(Sum('deduction_amount'))['deduction_amount__sum'] or 0,
		'sum_net_salary': details.aggregate(Sum('net_salary'))['net_salary__sum'] or 0,
	}

	details.aggregate = aggregate

	return render(request, 'payroll/payroll_detail.html', {
		'payroll': payroll,
		'details': details,
		'is_admin': is_admin
	})


@login_required
def payroll_employee_detail(request, payroll_id, detail_id):
	payroll = get_object_or_404(Payroll, pk=payroll_id)
	detail = get_object_or_404(PayrollDetail, pk=detail_id, payroll=payroll)

	# Kiểm tra quyền admin
	is_admin = request.user.is_superuser or request.user.groups.filter(name='HR Admin').exists()

	# Nếu không phải admin và bảng lương bị vô hiệu hóa, chuyển hướng về danh sách
	if not is_admin and payroll.status == 'disabled':
		messages.warning(request, "Bảng lương này đã bị vô hiệu hóa.")
		return redirect('payroll_list')

	# Lấy phụ cấp và khấu trừ
	allowances = detail.allowances.all()
	deductions = detail.deductions.all()

	return render(request, 'payroll/payroll_employee_detail.html', {
		'payroll': payroll,
		'detail': detail,
		'allowances': allowances,
		'deductions': deductions
	})


@login_required
def disable_payroll(request, pk):
	payroll = get_object_or_404(Payroll, pk=pk)

	# Kiểm tra quyền admin
	is_admin = request.user.is_superuser or request.user.groups.filter(name='HR Admin').exists()

	# Nếu không phải admin, không cho phép vô hiệu hóa
	if not is_admin and payroll.status == 'paid':
		messages.warning(request, "Bạn không có quyền vô hiệu hóa bảng lương đã thanh toán.")
		return redirect('payroll_detail', pk=pk)

	if request.method == 'POST':
		payroll.status = 'disabled'
		payroll.save()
		messages.success(request, f"Đã vô hiệu hóa bảng lương {payroll.name}")
		return redirect('payroll_detail', pk=pk)

	return render(request, 'payroll/disable_payroll.html', {
		'payroll': payroll
	})


@login_required
def activate_payroll(request, pk):
	payroll = get_object_or_404(Payroll, pk=pk)

	# Kiểm tra quyền admin
	is_admin = request.user.is_superuser or request.user.groups.filter(name='HR Admin').exists()

	# Nếu không phải admin, không cho phép kích hoạt
	if not is_admin:
		messages.warning(request, "Bạn không có quyền kích hoạt bảng lương.")
		return redirect('payroll_list')

	if request.method == 'POST':
		payroll.status = 'draft'
		payroll.save()
		messages.success(request, f"Đã kích hoạt bảng lương {payroll.name}")
		return redirect('payroll_detail', pk=pk)

	return render(request, 'payroll/activate_payroll.html', {
		'payroll': payroll
	})


@login_required
def export_payroll_excel(request, pk):
	payroll = get_object_or_404(Payroll, pk=pk)

	# Kiểm tra quyền admin
	is_admin = request.user.is_superuser or request.user.groups.filter(name='HR Admin').exists()

	# Nếu không phải admin và bảng lương bị vô hiệu hóa, chuyển hướng về danh sách
	if not is_admin and payroll.status == 'disabled':
		messages.warning(request, "Bảng lương này đã bị vô hiệu hóa.")
		return redirect('payroll_list')

	# Tạo workbook mới
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = f"Bảng lương {payroll.month}/{payroll.year}"

	# Định dạng
	title_font = Font(name='Arial', size=14, bold=True)
	header_font = Font(name='Arial', size=12, bold=True)
	normal_font = Font(name='Arial', size=11)

	# Tiêu đề
	ws.merge_cells('A1:J1')
	ws['A1'] = f"BẢNG LƯƠNG {payroll.position.name.upper()} - THÁNG {payroll.month}/{payroll.year}"
	ws['A1'].font = title_font
	ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

	# Thông tin bảng lương
	ws['A3'] = "Trạng thái:"
	ws['B3'] = dict(Payroll.STATUS_CHOICES)[payroll.status]
	ws['A4'] = "Ngày tạo:"
	ws['B4'] = payroll.created_at.strftime("%d/%m/%Y %H:%M")
	ws['A5'] = "Người tạo:"
	ws['B5'] = payroll.created_by.get_full_name() if payroll.created_by.get_full_name() else payroll.created_by.username

	# Header
	headers = [
		'STT', 'Mã NV', 'Họ và tên', 'Lương cơ bản', 'Công chuẩn', 'Công thực tế',
		'Tỷ lệ hưởng', 'Tổng thu nhập', 'Thuế TNCN', 'Kỷ luật', 'Khen thưởng', 'Khấu trừ khác', 'Thực lĩnh'
	]

	for col_num, header in enumerate(headers, 1):
		cell = ws.cell(row=7, column=col_num)
		cell.value = header
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

	# Dữ liệu
	details = payroll.payroll_details.all()
	row_num = 8

	for i, detail in enumerate(details, 1):
		ws.cell(row=row_num, column=1).value = i
		ws.cell(row=row_num, column=2).value = detail.employee.code
		ws.cell(row=row_num, column=3).value = detail.employee.full_name
		ws.cell(row=row_num, column=4).value = detail.basic_salary
		ws.cell(row=row_num, column=5).value = detail.standard_workdays
		ws.cell(row=row_num, column=6).value = detail.actual_workdays
		ws.cell(row=row_num, column=7).value = float(detail.attendance_ratio)
		ws.cell(row=row_num, column=8).value = detail.gross_salary
		ws.cell(row=row_num, column=9).value = detail.income_tax
		ws.cell(row=row_num, column=10).value = detail.discipline_amount
		ws.cell(row=row_num, column=11).value = detail.reward_amount
		ws.cell(row=row_num, column=12).value = detail.deduction_amount
		ws.cell(row=row_num, column=13).value = detail.net_salary

		row_num += 1

	# Tổng cộng
	ws.cell(row=row_num + 1, column=1).value = "Tổng cộng:"
	ws.merge_cells(f'A{row_num + 1}:G{row_num + 1}')
	ws.cell(row=row_num + 1, column=1).font = header_font
	ws.cell(row=row_num + 1, column=1).alignment = Alignment(horizontal='right')

	ws.cell(row=row_num + 1, column=8).value = sum(detail.gross_salary for detail in details)
	ws.cell(row=row_num + 1, column=9).value = sum(detail.income_tax for detail in details)
	ws.cell(row=row_num + 1, column=10).value = sum(detail.discipline_amount for detail in details)
	ws.cell(row=row_num + 1, column=11).value = sum(detail.reward_amount for detail in details)
	ws.cell(row=row_num + 1, column=12).value = sum(detail.deduction_amount for detail in details)
	ws.cell(row=row_num + 1, column=13).value = sum(detail.net_salary for detail in details)

	# Định dạng cột
	for col in range(1, 14):
		ws.column_dimensions[get_column_letter(col)].width = 15

	ws.column_dimensions['C'].width = 25

	# Định dạng số
	for row in range(8, row_num + 2):
		for col in [4, 8, 9, 10, 11, 12, 13]:
			cell = ws.cell(row=row, column=col)
			if cell.value is not None and isinstance(cell.value, (int, float, Decimal)):
				cell.number_format = '#,##0'

	# Tạo response
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename=Bang_luong_{payroll.month}_{payroll.year}.xlsx'

	wb.save(response)
	return response


@login_required
def approve_payroll(request, pk):
	payroll = get_object_or_404(Payroll, pk=pk)

	# Kiểm tra quyền admin
	is_admin = request.user.is_superuser or request.user.groups.filter(name='HR Admin').exists()

	# Nếu không phải admin và bảng lương không ở trạng thái nháp, không cho phép duyệt
	if not is_admin and payroll.status != 'draft':
		messages.warning(request, "Bạn không có quyền duyệt bảng lương này.")
		return redirect('payroll_detail', pk=pk)

	if request.method == 'POST':
		payroll.status = 'approved'
		payroll.save()
		messages.success(request, f"Đã duyệt bảng lương {payroll.name}")
		return redirect('payroll_detail', pk=pk)

	return render(request, 'payroll/approve_payroll.html', {
		'payroll': payroll
	})


@login_required
def mark_as_paid(request, pk):
	payroll = get_object_or_404(Payroll, pk=pk)

	# Kiểm tra quyền admin
	is_admin = request.user.is_superuser or request.user.groups.filter(name='HR Admin').exists()

	# Nếu không phải admin và bảng lương không ở trạng thái đã duyệt, không cho phép đánh dấu đã thanh toán
	if not is_admin and payroll.status != 'approved':
		messages.warning(request, "Bạn không có quyền đánh dấu bảng lương này đã thanh toán.")
		return redirect('payroll_detail', pk=pk)

	if request.method == 'POST':
		payroll.status = 'paid'
		payroll.save()
		messages.success(request, f"Đã đánh dấu bảng lương {payroll.name} đã thanh toán")
		return redirect('payroll_detail', pk=pk)

	return render(request, 'payroll/mark_as_paid.html', {
		'payroll': payroll
	})